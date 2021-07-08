
from openpyxl import load_workbook
from openpyxl import Workbook
import account_details


def access_file():
    return load_workbook(filename="Main_python_projrct.xlsx")


if __name__ == "__main__":
    Python_file = access_file().active
    for i in range(2, 17):
        print(Python_file.cell(row=i, column=1).value)
    PS_NO = int(input("Enter the PS No. of the candidate whose details you want to print: "))

    FILE_NAME = "Python_final_output.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "PS Number"
    sheet["A2"] = PS_NO

    index = account_details.index_of_ps_no(PS_NO)

    per = account_details.personal(index)
    ROW_NO = 2
    COL_NO = 2
    per1 = per[0]
    per2 = per[1]
    for i, value in enumerate(per1, start=ROW_NO):
        sheet.cell(row=i, column=COL_NO).value = value
    ROW_NO = 2
    COL_NO = 3
    for i, value in enumerate(per2, start=ROW_NO):
        sheet.cell(row=i, column=COL_NO).value = value

    LST_OP = account_details.school_mark_sheet(index)
    ROW_NO = 2
    COL_NO = 4
    for i, value in enumerate(LST_OP, start=ROW_NO):
        sheet.cell(row=i, column=COL_NO).value = value

    DOM_OP = account_details.domain_expertise(index)
    ROW_NO = 2
    COL_NO = 5
    for i, value in enumerate(DOM_OP, start=ROW_NO):
        sheet.cell(row=i, column=COL_NO).value = value

    hob = account_details.hobbies(index)
    ROW_NO = 2
    COL_NO = 6
    for i, value in enumerate(hob, start=ROW_NO):
        sheet.cell(row=i, column=COL_NO).value = value

    cit = account_details.cities(index)
    ROW_NO = 2
    COL_NO = 7
    for i, value in enumerate(cit, start=ROW_NO):
        sheet.cell(row=i, column=COL_NO).value = value

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 22.5
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 27
    sheet.column_dimensions['F'].width = 16
    sheet.column_dimensions['G'].width = 40

    workbook.save(filename=FILE_NAME)