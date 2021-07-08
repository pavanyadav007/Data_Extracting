from openpyxl import load_workbook


def index_of_ps_no(ps_no):
    ps_no_col = load_workbook(filename="Main_python_projrct.xlsx").active
    ind = int()
    for i in range(2, 17):
        if ps_no_col.cell(row=i, column=1).value == ps_no:
            ind = i
    return ind


def school_mark_sheet(index=1):
    mark_sheet = load_workbook(filename="Main_python_projrct.xlsx")
    marksheetopen = mark_sheet["Marksheet"]
    lst = []
    for i in range(2, 14):
        lst.append(marksheetopen.cell(row=index, column=i).value)

    return lst


def domain_expertise(index):
    domain = load_workbook(filename="Main_python_projrct.xlsx")
    domainopen = domain["Domain_Expertise"]
    dom = []
    for i in range(2, 16):
        if int(domainopen.cell(row=index, column=i).value) == 1:
            dom.append(domainopen.cell(row=1, column=i).value)
    return dom


def hobbies(index):
    hobby = load_workbook(filename="Main_python_projrct.xlsx")
    hobbyopen = hobby["Hobbies"]
    hobs = []
    for i in range(2, 15):
        if int(hobbyopen.cell(row=index, column=i).value) == 1:
            hobs.append(hobbyopen.cell(row=1, column=i).value)

    return hobs


def cities(index):
    city = load_workbook(filename="Main_python_projrct.xlsx")
    cityopen = city["Cities"]
    citlist = []
    for i in range(2, 16):
        if int(cityopen.cell(row=index, column=i).value) == 1:
            citlist.append(cityopen.cell(row=1, column=i).value)

    return citlist


def personal(index):
    personal_details = load_workbook(filename="Main_python_projrct.xlsx")
    personaldetailsopen = personal_details["Personal_Details"]
    fin_list = []
    lstinfo = []
    lsthead = []
    for i in range(2, 11):
        lstinfo.append(personaldetailsopen.cell(row=index, column=i).value)

    for i in range(2, 11):
        lsthead.append(personaldetailsopen.cell(row=1, column=i).value)
    fin_list.append(lsthead)
    fin_list.append(lstinfo)
    return fin_list
